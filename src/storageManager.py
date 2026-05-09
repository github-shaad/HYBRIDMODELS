"""
Simple Class to save data/models/statistics/plots.
"""
from config.config import *
import numpy as np
import openpyxl
from plots import *
import joblib
import json

class StorageManager:
    @staticmethod
    def store_predictions(data, prefix:str, modelspec:str):
        """
        Store Predictions \n
        prefix : Model or Portfolio \n
        modelspec : Which model's predictions being stored
        """
        predictions_path = PREDICTIONS_DIR / f"{prefix}_predictions" / modelspec
        np.save(predictions_path, data)
    
    @staticmethod
    def store_statistics(prefix, stat, model, value):
        path = None
        if prefix == "data":
            path = DATA_STATISTICS
        elif prefix == "model":
            path = MODEL_STATISTICS
        elif prefix == "portfolio":
            path = PORTFOLIO_STATISTICS
        else:
            raise ValueError("Wrong prefix type. Not data/model/portfolio")

        file_path = path / f"{prefix}_statistics.xlsx"
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        header_map = {cell.value: cell.column for cell in sheet[1] if cell.value is not None}



        if "Model" not in header_map or stat not in header_map:
            raise ValueError(f"Non-existent metrics or excel sheet structure.\nCheck excel file located at {file_path}")

        model_col = header_map["Model"]    
        stat_col = header_map[stat]

        target_row = None
        for row in range(2, sheet.max_row+1):
            if sheet.cell(row=row, column=model_col).value == model:
                target_row = row
                break

        if target_row is None:
            target_row = sheet.max_row + 1
            sheet.cell(row=target_row, column=model_col, value=model)

        sheet.cell(row=target_row, column=stat_col, value=value)

        workbook.save(file_path)
    
    @staticmethod
    def store_figures(plot, prefix, modelspec):
        path = FIGURES_DIR
        file_path = path / f"{prefix}_figures" / f"{modelspec}.png"
        plot.fig.savefig(file_path)
    @staticmethod
    def store_model(model, model_type, model_name):
        path = MODELS_DIR / model_type / f"{model_name}.joblib"
        joblib.dump(model, path)

    @staticmethod
    def store_params(params, model_name):
        path = PARAMS_DIR / f"{model_name}.json"
        
        # Open the file in write mode ('w')
        with open(path, 'w') as f:
            # Use dump (without the 's') to write to the file object
            json.dump(params, f, indent=4)
    
    @staticmethod
    def load_params(model_name):
        path = PARAMS_DIR / f"{model_name}.json"
        with open(path, "r") as f:
            params = json.load(f)

        for key, value in params.items():
            params[key] = [value]

        return params

class PlotSave:
    @staticmethod
    def multiPlot(tickers, true, pred, model_name, model_type):
        res_dict = {k:(t,p) for k,t,p in zip(tickers, true, pred)}
        m = MultiPlot(model_name)
        m.plot(res_dict, "True Prices", "Predicted Prices", figsize_per_plot=(10,6))
        StorageManager.store_figures(m, model_type, model_name)

    @staticmethod
    def tripleMultiPlot(tickers, true, base, pred, model_name, model_type):
        res_dict = {k:(t, b, p) for k,t,b,p in zip(tickers, true, base, pred)}
        m = TripleMultiPlot(model_name)
        m.plot(res_dict, "True Prices", "VARMAX Base Predictions", "ML Corrected Predictions",
               figsize_per_plot=(10,6))
        StorageManager.store_figures(m, model_type, f"{model_name}_comparison")